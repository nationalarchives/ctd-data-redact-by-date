# Function to read data from spreadsheet

import os, re, shutil, csv
from pprintpp import pprint as pp
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from datetime import date
from pathlib import Path

def getSpreadsheetValues(filename):
    ''' Gets spreadsheet by name and returns the spreadsheet as a worksheet and a list of column headings '''
    #path = os.path.join('data', filename) 
    wb = load_workbook(filename)
    
    sheet = wb.worksheets[0]
    values={}
    
    for col in sheet.columns:
        #column = [cell.value for cell in col if cell.value is not None]
        column = [cell.value if cell.value is not None else "" for cell in col]
        
        if len(column) > 0 and column.count("") != len(column):
            values[str(column[0]).strip()] = column[1:]
            
    return (values)

def getCoveringDatesbyPiece(filename):
    '''Gets a list of covering dates for each piece if there is a file with the specified name in the data/lib folder'''
    coveringDates = dict()

    if os.path.exists(os.path.join('data','lib', filename)):
        with open(os.path.join('data','lib', filename)) as dateFile:
            reader = csv.reader(dateFile, skipinitialspace=True)  
            coveringDates = {(int(line[0]) if line[0].isnumeric else line[0]):(int(line[1]) if line[1].isnumeric else line[1]) for line in reader}   
    else:
        print("No covering dates specified at " + os.path.join('lib', filename))
    
    return coveringDates

def removeBlanksFromColumn(column):
    return [value for value in column if value != ""]  

def getAgeFromColumn(column):
    ''' Get age from named column, if no age given then assume age is 18, and return a list of ages '''
    # if value is number then age otherwise default value      
    return [entry if str(entry).strip().isnumeric() else 18 for entry in removeBlanksFromColumn(column)]         

def getYearFromColumn(column, coveringDate):
    ''' Get year from named column and return a dictionary of years and covering dates'''
    # regex for dddd in text value
    # since Python 3.8 := allows you to name an evaluation as a variable which you can use int he list comparhension see https://stackoverflow.com/questions/26672532/how-to-set-local-variable-in-list-comprehension 
    years = [int(years[0]) if len(years := re.findall(r'\d{4}', entry)) == 1 else years for entry in removeBlanksFromColumn(column)] 
    #pp(years)
       
    return codifyYears(years, coveringDate)

def getDateFromList(dateList, earliest, latest, default, max=True):
    foundDate = -1
 
    for date in dateList:
        date = int(date)
        if (foundDate == -1 and foundDate > earliest and foundDate < latest) or (max and date < latest and date > foundDate) or (not(max) and date > earliest and date < foundDate):
            foundDate = date
            
    if foundDate == -1:
        foundDate = default
        
    return foundDate

def codifyYears(yearsList, coveringDate=""):
    defaultYear = coveringDate
    earliest = 1935
    latest = 1946
    codifiedYears = []
    coveringDates = []
       
    for year in yearsList:
        if type(year) is not int:
            if len(year) > 1:
                latestDate = "" 
                
                if int(max(year)) > earliest and int(max(year)) < latest:
                    latestDate = int(max(year))
                else:
                    latestDate = getDateFromList(year, earliest, latest, defaultYear)
                
                if int(min(year)) > earliest and int(min(year)) < latest:
                    earliestDate = int(min(year))
                else:
                    earliestDate = getDateFromList(year, earliest, latest, defaultYear, False)
                
                codifiedYears.append(latestDate)
                coveringDates.append(str(earliestDate) + " - " + str(latestDate))
            else:
                codifiedYears.append(defaultYear)
                coveringDates.append(defaultYear)
        elif year < earliest or year > latest:
            codifiedYears.append(defaultYear)
            coveringDates.append(defaultYear)           
        else:
            codifiedYears.append(year)
            coveringDates.append(year)
    
    #print(codifiedYears)
    #print(coveringDates)
    return zip(codifiedYears, coveringDates)
             
def insertCoveringDateValues(sheet, dateList):
    ''' Insert covering date values into dictionary of spreadsheet values '''
    sheet["Covering Dates"] = dateList
    return sheet

def openingCalculation(age, year):
    ''' return the year the record can be opened (100 years after birth) given an age in a given year'''
    if year != '':
        return (year - age) + 101 
    else:
        return '?'

def deceasedCheck(openingDate, additionalInfo):
    ''' return the default year for opening unless the word deceased appears in the additional info '''
    if "Deceased" in additionalInfo and openingDate > date.today().year:
        return date.today().year
    else:
        return openingDate

def createOpeningList(agesList, yearsList):   
    ''' return a list of years in which the record will be open given a list of ages and list of years'''
    return list(map(openingCalculation, agesList, yearsList))

def unredactIfDeceased(openingList, additionalInfoList):
    ''' update the opening list to current year if the person is dead'''
    return list(map(deceasedCheck, openingList, additionalInfoList))

def sheetRedactionNeededCheck(openingList):
    ''' return False if no opening dates in the given list are later than the current year'''
    return False if max(openingList) <= date.today().year else True
       
def selectByYear(previousRedactionList, currentRedactionList):
    ''' Return filter of changed redaction'''    
    return [False if a == b else True for a, b in zip(previousRedactionList, currentRedactionList)]    

def yearsToPublish(openingList, year=date.today().year):
    return list(range(year, max(openingList)+1))

def redactColumns(columnsToRedact, openingList, lastYearInSeries, year=date.today().year, minimum=True):
    ''' given a dictionary containung the columns that may need redacting, return a dict containing the original record values and
    the processed values for each year, by year, until all records have been opened. 
    
    {"base": [{col1Name:col1, col2Name:col2}], 
    year: [{filter: filter, col1Name: col1_redacted, col2Name: col2_redacted], 
    year+1 [{filter: filter, col1Name: col1_redacted, col2Name: col2_redacted]... 
    max_year: [{filter: filter, col1Name: col1_redacted, col2Name: col2_redacted]}
    
    '''
    
    boilerplate = "[Additional information regarding this case will be added to the catalogue when the case becomes over 100 years old. In cases when the date is not known, the latest date in the series (" + str(lastYearInSeries) + ") will be used]"
    
    processedColumns = {"base":columnsToRedact}
    
    yearsToPublishList = yearsToPublish(openingList)
    previousRedactions = []
    
    for currentYear in yearsToPublishList:
        #print(currentYear)
        
        toRedact = [True if currentYear < openingYear else False for openingYear in openingList]
        #test_redactByYear_testFile(toRedact, currentYear)
        #print(toRedact)
        
        filter = [True] * len(openingList)
        
        if previousRedactions == []:
            previousRedactions = toRedact
        else:
            filter = selectByYear(previousRedactions, toRedact)
            previousRedactions = toRedact
        
        #test_selectByYear_testFile(filter, currentYear)
        
        processedColumns[currentYear] = {"filter": filter}
    
        for columnName, column in columnsToRedact.items():
            newColumn = [boilerplate if record[1] and record[0] != "" else record[0] for record in zip(column, toRedact)]   
            processedColumns[currentYear][columnName]=newColumn
    
    return processedColumns

def pathToFile(year):
    path = os.path.join('data', 'converted',str(year))
    if not os.path.exists(path):
        os.makedirs(path)
    return path

def unredactByYear(filename, values, newValues, year, min=True):
    ''' print out a new spreadsheet with the full text for all columns for just the rows where the year is 100 years since birth'''
    
    wb = Workbook()
    newSheet = wb.active
    
    col = 1
    
    #print(values)
    
    #print(newValues[year].keys())
    
    filter = newValues[year]["filter"]
    
    for title, column in values.items():
        newSheet.cell(1, col, title).font = Font(bold=True)
    
        row = 2
        
        if title in newValues[year].keys():
            column = newValues[year][title]

        filteredColumn = zip(column, filter)           

        for filteredRow in filteredColumn:
            #print(str(row[1]) + ": " + str(x) + ", " + str(y))
            if (min and filteredRow[1]) or not min:
                newSheet.cell(row, col, filteredRow[0])
                row+=1
            
        col+=1 
    
    #print("Last cel written for " + str(year) + " x:" + str(row) + ", y: " + str(col))
    if row > 2:   
        path = pathToFile(year)  
        newFilename = os.path.splitext(os.path.basename(filename))[0] + "_" + str(year) + os.path.splitext(os.path.basename(filename))[1]
        newFile = os.path.join(path, newFilename)  
        wb.save(newFile)

def spreadsheetNoRedactions(filename, values):
    ''' print out a new spreadsheet with the full text for all columns'''

    wb = Workbook()
    newSheet = wb.active
    
    col = 1
    
    for title, column in values.items():
        newSheet.cell(1, col, title).font = Font(bold=True)
        row = 2
        for row_value in column:
            newSheet.cell(row, col, row_value)
            row+=1    
        col+=1 
    
    #print("Last cel written for " + str(year) + " x:" + str(row) + ", y: " + str(col))
    if row > 2:   
        path = pathToFile(date.today().year)  
        newFilename = os.path.splitext(os.path.basename(filename))[0] + '_NoRedactions' + os.path.splitext(os.path.basename(filename))[1]
        newFileFullPath = os.path.join(path, newFilename)  
        wb.save(newFileFullPath) 
        return newFileFullPath 

def generateSpreadsheets(filename, values, newValues, openingList):
    ''' Create the spreadsheets for each year '''
    yearsToPublishList = yearsToPublish(openingList)
    
    for currentYear in yearsToPublishList:
        unredactByYear(filename, values, newValues, currentYear)

def generateSummary(filename, ageList, coveringDatesList, openingListByPiece, openingListByExtractedDate, changesToOpening, full=False):
    ''' print out a summary spreadsheet, including covering dates comparison outcome, with information about each piece on a seperate tab '''

    if os.path.exists(os.path.join('data', 'summary', 'summary.xlsx')):
        wb = load_workbook(os.path.join('data', 'summary', 'summary.xlsx'))
        ws = wb.create_sheet()
    else:
        wb = Workbook()        
        ws = wb.active
        
    ws.title = filename
    colHeadings = ["Item", "Age", "Covering Dates", "Opening Year", "Opening Note"]
    
    col = 1
    count = 1
    
    for heading in colHeadings:
        row = 2
        ws.cell(1, col, heading).font = Font(bold=True)
        
        for age, coveringDate, openingByPiece, openingByExtractedDate, changeToOpening in zip(ageList, coveringDatesList, openingListByPiece, openingListByExtractedDate, changesToOpening):
            ws.cell(row, 1, row-1)
            ws.cell(row, 2, age)
            ws.cell(row, 3, coveringDate)
            ws.cell(row, 4, openingByExtractedDate)
            if openingByExtractedDate != '?' and openingByPiece <= date.today().year and openingByExtractedDate <= date.today().year and full:
                ws.cell(row, 5, "Difference in opening date irrelevant because both before current year. " + changeToOpening)
            elif openingByExtractedDate != '?' and openingByExtractedDate < openingByPiece and full:
                ws.cell(row, 5, "Date in description earlier than supplied covering date. Earlier date of " + str(openingByExtractedDate) + " used. " + changeToOpening)
                if '!' not in ws.title:
                    ws.title += '!'
            elif openingByExtractedDate != '?' and openingByExtractedDate > openingByPiece:
                ws.cell(row, 5, "Date in description later than supplied covering date. Later date of " + str(openingByExtractedDate) + " used. Covering date should be checked! " + changeToOpening)
                if '!' not in ws.title:
                    ws.title += '!'
            elif changeToOpening != "":
                ws.cell(row, 5, changeToOpening)
            elif openingByExtractedDate == '?':
                ws.cell(row, 5, "No opening date found")

            row += 1
            
        col += 1
            
    wb.save(os.path.join('data', 'summary', 'summary.xlsx'))       
    
        
def getFileList(myDir):
    ''' Get a list of xlsx files in the given directory '''
    return [file for file in myDir.glob("[!~.]*.xlsx")]

def generateFiles(coveringDateFile='',reset=True,output=True,summary=True):
    ''' Main program. Expects spreadsheets to be in the data directory.'''
    
    if reset and os.path.exists(os.path.join('data', 'converted')):        
        shutil.rmtree(os.path.join('data', 'converted'))
        os.makedirs(os.path.join('data', 'converted'))
        
    if summary:
        if os.path.exists(os.path.join('data', 'summary', 'summary.xlsx')):
            os.remove(os.path.join('data', 'summary', 'summary.xlsx'))
    
    for file in getFileList(Path('data')):
        print("Processing " + os.path.basename(file))
        currentSpreadsheet = getSpreadsheetValues(file)
        #print(currentSpreadsheet.keys())
        
        try:        
            test_load_file(list(currentSpreadsheet.keys()))
        
            ageList = getAgeFromColumn(currentSpreadsheet['Age'])
            test_all_ints(ageList)

            if '.' in coveringDateFile:
                #filename = os.path.splitext(os.path.basename(file))[0]
                piece = os.path.splitext(os.path.basename(file))[0].split('_')[1]
                getCoveringDatesbyPieceDict = getCoveringDatesbyPiece(coveringDateFile)

                coveringDatebyPiece = ""
                if len(getCoveringDatesbyPieceDict) > 0:
                    coveringDatebyPiece = getCoveringDatesbyPieceDict[int(piece)]

        
            dates = getYearFromColumn(currentSpreadsheet['Brief summary of grounds for recommendation'], coveringDatebyPiece)        
                
            yearList = []
            coveringDatesList = []
            coveringDatesByPieceList = []

            for parts in dates:
                #print(parts)
                yearList.append(parts[0])
                coveringDatesList.append(parts[1])
                coveringDatesByPieceList.append(coveringDatebyPiece)

            '''
            if yearList != coveringDatesByPieceList:
                print("Possibly anomologus dates in spreadsheet found (expected " + str(coveringDatebyPiece) + "): ", end="")
                otherYears = set(coveringDatesList)
                if coveringDatebyPiece in otherYears:
                    otherYears.remove(coveringDatebyPiece)
                print(otherYears)
            '''                
            
            insertCoveringDateValues(currentSpreadsheet, coveringDatesByPieceList)
            print("Adding covering dates for " + os.path.basename(file))

            originalOpeningListByPiece = createOpeningList(ageList, coveringDatesByPieceList)
            originalOpeningListByExtractedDate = createOpeningList(ageList, yearList)

            additionalInfoList = currentSpreadsheet['Additional Information']

            openingListByPiece = unredactIfDeceased(originalOpeningListByPiece, additionalInfoList)
            openingListByExtractedDate = unredactIfDeceased(originalOpeningListByExtractedDate, additionalInfoList)

            changesToOpening = test_unredaction_due_to_death(originalOpeningListByExtractedDate, openingListByExtractedDate, additionalInfoList)
            #print(changesToOpening)
            '''
            combinedLists = zip(openingList, altOpeningList)

            row = 2
            for default, byRow in combinedLists:        
                if default > byRow and byRow > 2022:
                    print("Row " + str(row) + ": Opening date is " + str(default) + " but earlier date of " + str(byRow) + " might be possible")
                row += 1
            '''

            test_all_ints(openingListByExtractedDate)
            
        except AssertionError as e:
            print("Issue with " + os.path.basename(file) + " skipping")
            print(e)
            continue
        
    
        if output:
            if(sheetRedactionNeededCheck(openingListByExtractedDate)):
                newColumnValues = redactColumns(dict((key, currentSpreadsheet[key]) for key in ['Occupation', 'Brief summary of grounds for recommendation']), openingListByExtractedDate, 1946)
                generateSpreadsheets(os.path.basename(file), currentSpreadsheet, newColumnValues, openingListByExtractedDate)
                
                newFilename = os.path.splitext(os.path.basename(file))[0] + "_" + str(date.today().year) + os.path.splitext(os.path.basename(file))[1]
                pathToNewFile = os.path.join(pathToFile(date.today().year), newFilename)

                print(os.path.basename(file) + " redacted. Spreadsheets with redacted descriptions and unredactions generated.")
            else:
                path = pathToFile(date.today().year)
                pathToNewFile = spreadsheetNoRedactions(os.path.basename(file), currentSpreadsheet)

                '''
                filename = os.path.splitext(os.path.basename(file))[0] + '_NoRedactions' + os.path.splitext(os.path.basename(file))[1]
                
                try: 
                    shutil.copyfile(file, os.path.join(path, filename))
                except shutil.SameFileError:
                    pass
                '''

                print(os.path.basename(file) + " copied over, no redactions needed")

            test_load_file_row_count(removeBlanksFromColumn(currentSpreadsheet['Item']), pathToNewFile)


        if summary:   
            generateSummary(os.path.splitext(os.path.basename(file))[0], ageList, coveringDatesByPieceList, openingListByPiece, openingListByExtractedDate, changesToOpening)   
            

      
        

### Tests ####


def test_load_file(columnHeadings):
    expectedColumns = ['Letter','Series','Piece', 'Item', 'Treasury case number', 'Home Office case number', 'First names/Initials', 'Surname', 'Age', 'Occupation', 'Award granted', 'Brief summary of grounds for recommendation', 'Additional Information']   
    assert columnHeadings == expectedColumns, "Error in expected columns. Check for " + str([i for i in expectedColumns + columnHeadings if i not in expectedColumns or i not in columnHeadings])  

def test_load_generated_file(columnHeadings):
    expectedColumns = ['Letter','Series','Piece', 'Item', 'Treasury case number', 'Home Office case number', 'First names/Initials', 'Surname', 'Age', 'Occupation', 'Award granted', 'Brief summary of grounds for recommendation', 'Additional Information', 'Covering Dates']   
    assert columnHeadings == expectedColumns, "Error in expected columns. Check for " + str([i for i in expectedColumns + columnHeadings if i not in expectedColumns or i not in columnHeadings])  

def test_load_file_row_count(item_column_on_load, newFile):
    newSpreadsheet = getSpreadsheetValues(newFile)
    test_load_generated_file(list(newSpreadsheet.keys()))
    assert all(item in item_column_on_load for item in removeBlanksFromColumn(newSpreadsheet['Item'])), "Error in expected output. Missing items: " + str([i for i in item_column_on_load if i not in newSpreadsheet['Item']])  

def test_all_ints(list):
    assert all(isinstance(x, int) for x in list), "Error in expected data types: " + str(list)

def test_get_covering_dates(coveringDateList):
    assert coveringDateList[2] == 1940, "Error in returned value: expected 1940 got " + coveringDateList[2]

def test_unredaction_due_to_death():
    openingList = [date.today().year, date.today().year+1, date.today().year+2, date.today().year+1, date.today().year]
    additionalInfoList = ["","Random other text","Deceased","Deceased",""]

    assert unredactIfDeceased(openingList, additionalInfoList) == [date.today().year, date.today().year+1, date.today().year, date.today().year, date.today().year]
    #assert unredactBecauseDeceased(openingList, additionalInfoList) == [date.today().year, date.today().year+1, date.today().year+2, date.today().year+1, date.today().year]

def test_unredaction_due_to_death(originalList, redactedList, spreadsheetColumn):
    changesToOpening = []
    for origDate, newDate, notes in zip(originalList, redactedList, spreadsheetColumn):
        if origDate != newDate:
            assert 'Deceased' in notes, "Error in opening date: A change has been made to the opening date but 'Deceased' not detected in Additional Notes"
            changesToOpening.append("Record opened because marked as deceased")
        elif 'Deceased' in notes:
            assert origDate <= date.today().year and newDate <= date.today().year, "Error in opening date: 'Deceased' detected in Additional Notes but opening date not set to current year"
            changesToOpening.append("Marked as deceased but record already open")
        else:
            changesToOpening.append("")
    
    return changesToOpening

'''    
def test_age_testFile(ageList):
    expectedAges = [16,18,18,20,25,45,16,18,20,16,18,20,16,18,18,20,16,18,20,36,16,18,20,30,25,23]   
    assert expectedAges == ageList

def test_year_testFile(yearList):
    expectedYears = [1936,1938,1938,1940,1945,1945,1937,1939,1941,1938,1940,1942,1938,1940,1941,1943,1940,1942,1944,1940,1941,1943,1945,1944,1943,1942]    
    assert expectedYears == yearList

def test_openingList_testFile(openingList):
    expectedOpening = [2020,2020,2020,2020,2020,2000,2021,2021,2021,2022,2022,2022,2022,2022,2023,2023,2024,2024,2024,2004,2025,2025,2025,2014,2018,2019]    
    assert expectedOpening == openingList 

def test_redactByYear_testFile(toRedactList, year):    
    if year == 2022:
        expectedRedactionList = [False,False,False,False,False,False,False,False,False,False,False,False,False,False,True,True,True,True,True,False,True,True,True,False,False,False]
        assert expectedRedactionList == toRedactList
    elif year == 2023:
        expectedRedactionList = [False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,True,True,True,False,True,True,True,False,False,False]
        assert expectedRedactionList == toRedactList
    elif year == 2024:
        expectedRedactionList = [False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,True,True,True,False,False,False]
        assert expectedRedactionList == toRedactList
    else:
        expectedRedactionList = [False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,False]
        assert expectedRedactionList == toRedactList

def test_selectByYear_testFile(selectList, year):    
    if year == 2022:
        expectedSelectList = [True,True,True,True,True,True,True,True,True,True,True,True,True,True,True,True,True,True,True,True,True,True,True,True,True,True]
        assert expectedSelectList == selectList
    elif year == 2023:
        expectedSelectList = [False,False,False,False,False,False,False,False,False,False,False,False,False,False,True,True,False,False,False,False,False,False,False,False,False,False]
        assert expectedSelectList == selectList
    elif year == 2024:
        expectedSelectList = [False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,True,True,True,False,False,False,False,False,False,False]
        assert expectedSelectList == selectList
    else:
        expectedSelectList = [False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,False,True,True,True,False,False,False]
        assert expectedSelectList == selectList
'''


        

### Main        

'''
testSpreadsheet = getSpreadsheetValues('test.xlsx')
test_loadfile(list(testSpreadsheet.keys()))

ageList = getAgeFromColumn(testSpreadsheet['Age'])
test_all_ints(ageList)
test_age_testFile(ageList)

yearList = getYearFromColumn(testSpreadsheet['Brief summary of grounds for recommendation'])
test_all_ints(yearList)
test_year_testFile(yearList)

insertCoveringDateValues(testSpreadsheet, yearList)

openingList = createOpeningList(ageList, yearList)
test_all_ints(openingList)
test_openingList_testFile(openingList)

if(sheetRedactionNeededCheck(openingList)):
    newColumns = redactColumns(dict((key, currentSpreadsheet[key]) for key in ['Occupation', 'Brief summary of grounds for recommendation']), openingList, 1945)
    #pp(newColumns)
    
#generateSpreadsheets("test.xlsx", currentSpreadsheet, newColumns, openingList)

test_unredactionDueToDeath()
'''

generateFiles('covering_dates.csv')

#test_get_covering_dates(getCoveringDatesbyPiece('covering_dates.csv'))